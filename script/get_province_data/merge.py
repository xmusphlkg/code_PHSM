import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from tqdm import tqdm
import time
import shutil
from dataclean import find_missing_elements

# The code is iterating over a list of provinces and for each province, it reads a CSV file containing
# disease data. It then creates a DataFrame with the data and saves it to an Excel file. The Excel
# file is used as a data center for storing disease data for different provinces. The code also
# removes any existing sheets with the same name in the Excel file before adding the new data.
provinces = os.listdir('./data/province')
data_center = pd.read_csv('./data/latest_data_center.csv')
book = load_workbook('./data/nation_and_provinces.xlsx')

for province in provinces:
    sheet_name = province
    if sheet_name in book.sheetnames:
        std = book[sheet_name]
        book.remove(std)
    data_list = []
    df_report = pd.read_csv(f'./data/province/{province}/{province}.csv', encoding='gbk')
    for i in range(len(df_report)):
        data_list.append([pd.to_datetime(f"{df_report['year'][i]}-{df_report['month'][i]:02}", format='%Y-%m'),
                          df_report['发病数'][i], re.sub('[^\u4e00-\u9fa5a-zA-Z]', '', df_report['疾病病种'][i]), None,
                          'Report', df_report['url'][i],
                          df_report['year'][i], df_report['month'][i]])
    df_list = pd.DataFrame(data_list,
                           columns=["date", "value", "disease_cn", "disease_en", "source", "url", "year", "month"])
    print(province)
    try:
        with pd.ExcelWriter('./data/nation_and_provinces.xlsx', engine='openpyxl') as writer:
            book = writer.book
            df_list.to_excel(writer, sheet_name=sheet_name, index=False)
    except:
        pass
book.save('./data/nation_and_provinces.xlsx')


def capitalize_sheet_names(file_path):
    """
    The function `capitalize_sheet_names` takes a file path as input, loads an Excel workbook, and
    capitalizes the first letter of each sheet name while removing any non-Chinese or non-alphabetic
    characters.
    
    :param file_path: The file path is the location of the Excel file that you want to modify. In this
    case, the file path is './data/nation_and_provinces.xlsx'
    """
    workbook = load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        first_letter = sheet_name[0].upper()
        new_sheet_name = first_letter + sheet_name[1:]
        workbook[sheet_name].title = re.sub('[^\u4e00-\u9fa5a-zA-Z]', '', new_sheet_name)
    workbook.save(file_path)

file_path = './data/nation_and_provinces.xlsx'
capitalize_sheet_names(file_path)

# The code is creating a list called `no_str_list` to store disease names that are not present in the
# `DiseasesCN` column of the `data_center` DataFrame.
no_str_list = []
for sheet_name in book.sheetnames:
    df_sheet = pd.read_excel(file_path, sheet_name)
    for i in tqdm(range(len(df_sheet)), desc="Processing", unit="iteration"):
        if 'disease_cn' in df_sheet.columns:
            disease_cn_value = df_sheet['disease_cn'][i]
            if disease_cn_value not in data_center['DiseasesCN'].values:
                no_str_list.append(disease_cn_value)
no_str_list = list(set(no_str_list))
no_str_list = pd.DataFrame(no_str_list, columns=['disease_cn'])
diseaseName2Code = pd.read_csv('./liu_script/diseaseName2Code.csv', encoding='gbk')
no_str_list = find_missing_elements(no_str_list, 'disease_cn', diseaseName2Code, 'Name')
no_str_list = pd.DataFrame(no_str_list, columns=['disease_cn'])
no_str_list.to_csv('./liu_script/diseaseName2diseaseNameCN.csv', encoding='gbk', index=False)


# The above code is processing data from a data center and saving it to an Excel file. It replaces the
# value 'Total' in the 'Province' column with 'Nation', sorts the data by 'Province', and resets the
# index. Then, it iterates over the data and creates a new sheet in the Excel file for each province.
# It appends the data for each province to a list, and when it reaches the last row for a province or
# encounters a different province, it creates a DataFrame from the list and saves it to the Excel
# file. Finally, it makes a copy of the Excel
file_path = './data/nation_and_provinces.xlsx'
origin_file_path = './data/origin_nation_and_provinces.xlsx'
data_center['Province'].replace('Total', 'Nation', inplace=True)
data_center.sort_values(by=['Province'], inplace=True)
data_center.reset_index(drop=True, inplace=True)
data_list = []
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    for i in tqdm(range(len(data_center)), desc="Processing", unit="iteration"):
        sheet_name = data_center['Province'][i]
        if i == len(data_center) - 1:
            data_list.append([data_center['Date'][i], data_center['Cases'][i], data_center['DiseasesCN'][i],
                              data_center['Diseases'][i], 'DataCenter', data_center['URL'][i],
                              str(data_center['YearMonthDay'][i]).split('/')[0],
                              str(data_center['YearMonthDay'][i]).split('/')[1]])
            df_list = pd.DataFrame(data_list,
                                   columns=["date", "value", "disease_cn", "disease_en", "source", "url", "year",
                                            "month"])
            try:
                existing_df = pd.read_excel(origin_file_path, sheet_name=sheet_name)
                combined_df = pd.concat([existing_df, df_list], axis=0)
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except:
                df_list.to_excel(writer, sheet_name=sheet_name, index=False)
            data_list = []
        elif data_center['Province'][i+1] != data_center['Province'][i] :
            data_list.append([data_center['Date'][i], data_center['Cases'][i], data_center['DiseasesCN'][i],
                              data_center['Diseases'][i], 'DataCenter', data_center['URL'][i],
                              str(data_center['YearMonthDay'][i]).split('/')[0],
                              str(data_center['YearMonthDay'][i]).split('/')[1]])
            df_list = pd.DataFrame(data_list,
                                   columns=["date", "value", "disease_cn", "disease_en", "source", "url", "year",
                                            "month"])
            try:
                existing_df = pd.read_excel(origin_file_path, sheet_name=sheet_name)
                combined_df = pd.concat([existing_df, df_list], axis=0)
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
            except:
                df_list.to_excel(writer, sheet_name=sheet_name, index=False)
            data_list = []
        else:
            data_list.append([data_center['Date'][i], data_center['Cases'][i], data_center['DiseasesCN'][i],
                              data_center['Diseases'][i], 'DataCenter', data_center['URL'][i],
                              str(data_center['YearMonthDay'][i]).split('/')[0],
                              str(data_center['YearMonthDay'][i]).split('/')[1]])
shutil.copyfile('./data/nation_and_provinces.xlsx', './data/nation_and_provinces_1.xlsx')

# The code block is opening an Excel file using `pd.ExcelWriter` and the `openpyxl` engine. It then
# iterates over each sheet in the Excel file and performs the following steps:
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    for sheet_name in writer.book.sheetnames:
        try:
            df_sheet = pd.read_excel("./data/nation_and_provinces_1.xlsx", sheet_name)
            df_sheet.sort_values(by=['date'], inplace=True)
            df_sheet.reset_index(drop=True, inplace=True)
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        except:
            pass

#Build Chinese-Chinese dictionary and Chinese-English dictionary
file_path = './data/nation_and_provinces.xlsx'
shutil.copyfile('./data/nation_and_provinces.xlsx', './data/nation_and_provinces_1.xlsx')
dict_cn_df =pd.read_csv('./liu_script/diseaseName2diseaseNameCN.csv',encoding='gbk')
dict_df=pd.read_csv('./liu_script/dict.csv',encoding='gbk')
dict={}
dict_cn={}
for i in range(len(dict_cn_df)):
    dict_cn.update({dict_cn_df['disease_cn'][i]:dict_cn_df['Name'][i]})
for i in range(len(dict_df)):
    dict.update({dict_df['Unnamed: 0'][i]:dict_df['0'][i]})

#Dictionary matching
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    for sheet_name in writer.book.sheetnames[4:]:
        df_sheet = pd.read_excel("./data/nation_and_provinces_1.xlsx", sheet_name)
        for i in tqdm(range(len(df_sheet)), desc="Processing", unit="iteration"):
            disease_cn_value = df_sheet['disease_cn'][i]
            if disease_cn_value in dict_cn:
                disease_cn_value_update = dict_cn[disease_cn_value]
                df_sheet['disease_cn'][i] = disease_cn_value_update
        df_sheet.to_excel(writer, sheet_name, index=False)
shutil.copyfile('./data/nation_and_provinces.xlsx', './data/nation_and_provinces_1.xlsx')

# The code block is opening an Excel file using `pd.ExcelWriter` and the `openpyxl` engine. It then
# iterates over each sheet in the Excel file (starting from the 5th sheet) and performs the following
# steps:
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    for sheet_name in writer.book.sheetnames[4:]:
        df_sheet = pd.read_excel("./data/nation_and_provinces_1.xlsx", sheet_name)
        for i in tqdm(range(len(df_sheet)), desc="Processing", unit="iteration"):
            disease_cn_value = df_sheet['disease_cn'][i]
            if disease_cn_value in dict:
                disease_en_value = dict[disease_cn_value]
                df_sheet['disease_en'][i] = disease_en_value
        df_sheet.to_excel(writer, sheet_name, index=False)
shutil.copyfile('./data/nation_and_provinces.xlsx', './data/nation_and_provinces_1.xlsx')

file_path = './data/nation_and_provinces.xlsx'
book = load_workbook('./data/nation_and_provinces.xlsx')
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    for sheet_name in writer.book.sheetnames[4:]:
        df_sheet = pd.read_excel("./data/nation_and_provinces_1.xlsx", sheet_name)
        df_sheet.dropna(inplace=True)
        df_sheet.to_excel(writer, sheet_name, index=False)
shutil.copyfile('./data/nation_and_provinces.xlsx', './data/nation_and_provinces_1.xlsx')